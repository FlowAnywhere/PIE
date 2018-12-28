from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import collections
import math
import multiprocessing
import os
import random
import re

import tensorflow as tf
from openpyxl import load_workbook

from PIE.config import Config


class DataProcessor(object):
    def __init__(self, config):
        self.config = config

    def tokenize_text(self, text):
        def chunks(l, n):
            """Yield successive n-sized chunks from l."""
            for i in range(0, len(l), n):
                yield l[i:i + n]

        token = self.config.tokenizer.tokenize(text)
        input_ids = self.config.tokenizer.convert_tokens_to_ids(token)

        return chunks(input_ids, self.config.max_seq_length)

    def get_dataset(self, is_training):
        name_to_features = {
            "input_ids": tf.FixedLenFeature([self.config.max_seq_length], tf.int64),
            "input_mask": tf.FixedLenFeature([self.config.max_seq_length], tf.int64),
            "segment_ids": tf.FixedLenFeature([self.config.max_seq_length], tf.int64),
            "label_ids": tf.FixedLenFeature([self.config.max_seq_length], tf.int64),
        }

        def decode_record(record, name_to_features):
            example = tf.parse_single_example(record, name_to_features)
            for name in list(example.keys()):
                t = example[name]
                if t.dtype == tf.int64:
                    t = tf.to_int32(t)
                example[name] = t
            return example

        tfrecord_files = []
        for root, dirs, files in os.walk(
                self.config.dataset_dir_train if is_training else self.config.dataset_dir_valid):
            for file in files:
                if file.endswith(".tfrecords"):
                    tfrecord_files.append(os.path.join(root, file))

        d = tf.data.TFRecordDataset(tfrecord_files, buffer_size=64000, num_parallel_reads=multiprocessing.cpu_count())

        if is_training:
            d = d.shuffle(buffer_size=100)

        return d.apply(
            tf.contrib.data.map_and_batch(
                lambda record: decode_record(record, name_to_features),
                batch_size=self.config.batch_size,
                drop_remainder=True
            ))

    def convert_to_tsv(self, original_file, new_file1, new_file2=None, split=0.0):
        raise NotImplementedError()

    def convert_to_tfrecord(self, tsv_file, tfrecord_file):
        def create_int_feature(values):
            f = tf.train.Feature(int64_list=tf.train.Int64List(value=list(values)))
            return f

        def convert_single_example(example):
            label_map = {}
            for (i, label) in enumerate(self.config.label_list, 1):
                label_map[label] = i

            textlist = example[0].split(' ')
            labellist = example[1].split(' ')
            tokens = []
            labels = []
            for i, word in enumerate(textlist):
                token = self.config.tokenizer.tokenize(word)
                tokens.extend(token)
                label_1 = labellist[i]
                for m in range(len(token)):
                    if m == 0:
                        labels.append(label_1)
                    else:
                        labels.append("X")

            if len(tokens) > self.config.max_seq_length - 2:
                tokens = tokens[0:(self.config.max_seq_length - 2)]
                labels = labels[0:(self.config.max_seq_length - 2)]
            ntokens = []
            segment_ids = []
            label_ids = []
            ntokens.append("[CLS]")
            segment_ids.append(0)
            # append("O") or append("[CLS]") not sure!
            label_ids.append(label_map["[CLS]"])
            for i, token in enumerate(tokens):
                ntokens.append(token)
                segment_ids.append(0)
                label_ids.append(label_map[labels[i]])
            ntokens.append("[SEP]")
            segment_ids.append(0)
            # append("O") or append("[SEP]") not sure!
            label_ids.append(label_map["[SEP]"])
            input_ids = self.config.tokenizer.convert_tokens_to_ids(ntokens)
            input_mask = [1] * len(input_ids)

            while len(input_ids) < self.config.max_seq_length:
                input_ids.append(0)
                input_mask.append(0)
                segment_ids.append(0)
                label_ids.append(0)
                ntokens.append("[PAD]")

            assert len(input_ids) == self.config.max_seq_length
            assert len(input_mask) == self.config.max_seq_length
            assert len(segment_ids) == self.config.max_seq_length
            assert len(label_ids) == self.config.max_seq_length

            return input_ids, input_mask, segment_ids, label_ids

        tf.gfile.MakeDirs(os.path.dirname(tfrecord_file))
        writer = tf.python_io.TFRecordWriter(tfrecord_file)

        example = []
        with open(tsv_file, mode='r', encoding='UTF-8') as f:
            for idx, line in enumerate(f, 1):
                example.append(line.strip())

                if idx % 2 == 0:
                    input_ids, input_mask, segment_ids, label_ids = convert_single_example(example)

                    features = collections.OrderedDict()
                    features["input_ids"] = create_int_feature(input_ids)
                    features["input_mask"] = create_int_feature(input_mask)
                    features["segment_ids"] = create_int_feature(segment_ids)
                    features["label_ids"] = create_int_feature(label_ids)

                    tf_example = tf.train.Example(features=tf.train.Features(feature=features))
                    writer.write(tf_example.SerializeToString())

                    example.clear()

        writer.close()


class ConllDataProcessor(DataProcessor):
    def convert_to_tsv(self, original_file, new_file1, new_file2=None, split=0.0):
        token_counter = {}

        tf.gfile.MakeDirs(os.path.dirname(new_file1))

        new_file_content = []

        words, labels = [], []
        with open(original_file, encoding='UTF-8') as f:
            for line in f:
                line = line.strip()

                if len(line) == 0 or line.startswith('-DOCSTART'):
                    if len(words) > 0:
                        word_line = ' '.join(words)
                        label_line = ' '.join(labels)
                        new_file_content.append(word_line)
                        new_file_content.append(label_line)

                        words, labels = [], []
                else:
                    line = line.split(' ')
                    words.append(line[0])

                    label = line[-1].split('-')
                    new_label = 'O'
                    if len(label) == 2 and label[1] == 'PER':
                        new_label = label[0] + '-' + 'PERSON'

                        # count
                        if label[0] == 'B':
                            if not token_counter.__contains__('PERSON'):
                                token_counter['PERSON'] = 0
                            token_counter['PERSON'] = token_counter['PERSON'] + 1

                    labels.append(new_label)

        with open(new_file1, mode="w", encoding='UTF-8') as f:
            for i, line in enumerate(new_file_content):
                f.write("{}\n".format(line))

        return token_counter


class XlsxDataProcessor(DataProcessor):
    def convert_to_tsv(self, original_file, new_file1, new_file2, split=0.5):
        def get_disabled_col():
            return ['description', 'areasofspecialization', 'professionalassociation', 'LocationInstructions',
                    'faxnumber',
                    'fax', 'phone1', 'phone2', 'phonefax', 'phonetollfree', 'telephonenumber', 'principalfirstname',
                    'principallastname', 'firstname', 'middlename', 'lastname', 'contact-lastname', 'contact-firstname']

        def get_tag(field_name, word_list_length, word_index, stat):
            tag1 = get_tag1(word_list_length, word_index)
            if field_name.lower() in ['fullname', 'contactname', 'principalname']:
                # if field_name.lower() in ['fullname', 'firstname', 'middlename', 'lastname', 'contact-lastname',
                #                           'contact-firstname', 'contactname', 'principalname', 'principalfirstname',
                #                           'principallastname', 'fundcontact']:
                if tag1 in ['B', 'S']:
                    if not stat.__contains__('PERSON'):
                        stat['PERSON'] = 0
                    stat['PERSON'] += 1
                return tag1 + '-' + 'PERSON'
            elif field_name.lower() in ['businessname', 'organization', 'localname', 'company', 'schoolname']:
                if tag1 in ['B', 'S']:
                    if not stat.__contains__('ORG'):
                        stat['ORG'] = 0
                    stat['ORG'] += 1
                return tag1 + '-' + 'ORG'
            # elif field_name.lower() in ['telephonenumber', 'telephone', 'phone', 'phonenumber', 'faxnumber', 'fax', 'phone1',
            #                             'phone2', 'phonefax', 'phonetollfree']:

            # elif field_name.lower() in ['telephone', 'phone', 'phonenumber']:  # , 'fax'
            #     if tag1 in ['B', 'S']:
            #         stat.PHONE += 1
            #     return tag1 + '-' + 'PHONE'
            # elif field_name.lower() in ['emailaddress', 'email', 'e-mail']:
            #     if tag1 in ['B', 'S']:
            #         stat.EMAIL += 1
            #     return tag1 + '-' + 'EMAIL'
            elif field_name.lower() in ['streetaddress', 'address', 'street', 'physicaladdress', 'buildingaddress']:
                if tag1 in ['B', 'S']:
                    if not stat.__contains__('ADDRESS'):
                        stat['ADDRESS'] = 0
                    stat['ADDRESS'] += 1
                return tag1 + '-' + 'ADDRESS'
            # elif field_name.lower() in ['postalcode']:
            #     return get_tag1(word_list_length, word_index) + '-' + 'POSTALCODE'
            else:
                return 'O'

        def get_tag1(word_list_length, word_index):
            return 'B' if word_index == 0 else 'I'

        token_counter1 = {}
        token_counter2 = {}

        wb = load_workbook(original_file, guess_types=False)
        split_idx = math.floor(wb.active.max_row * split)

        tf.gfile.MakeDirs(os.path.dirname(new_file1))
        tf.gfile.MakeDirs(os.path.dirname(new_file2))

        with open(new_file1, mode='w', encoding='UTF-8') as train_f:
            with open(new_file2, mode='w', encoding='UTF-8') as valid_f:

                header = None
                header_mask = None
                for idx, row in enumerate(wb.active.values):
                    if header is None:
                        header = [x.replace(' ', '') for x in row if x is not None]
                        header_mask = [False if x.lower() in get_disabled_col() else True for x in header]
                        continue

                    f = train_f if idx < split_idx else valid_f
                    stat = token_counter1 if idx < split_idx else token_counter2
                    row_string = []
                    for i, cell in enumerate(row):
                        if i < len(header):
                            row_string.append(str(cell).strip())

                    words, labels = [], []

                    for i, cell in enumerate(row):
                        if i < len(header) and header_mask[i]:

                            word_raw = re.split('[\s+\n]', row_string[i])
                            for j, word in enumerate(word_raw):
                                words.append(word)
                                labels.append('O' if cell is None or str(cell).lower() in ['na',
                                                                                           'not applicable'] else get_tag(
                                    header[i], len(word_raw), j, stat))

                            # separated by column
                            if random.randint(0, 100) > 50:
                                f.write(' '.join(words))
                                f.write('\n')
                                f.write(' '.join(labels))
                                f.write('\n')

                                words.clear()
                                labels.clear()

                    f.write(' '.join(words))
                    f.write('\n')
                    f.write(' '.join(labels))
                    f.write('\n')

                    words.clear()
                    labels.clear()

        return token_counter1, token_counter2


if __name__ == '__main__':
    from collections import Counter

    conll2003 = ConllDataProcessor(Config())
    # conll2003.tokenize_text('info@BDAccelerate.com')
    token_counter_train = conll2003.convert_to_tsv('../data/raw/conll2003/en/train_bio.txt',
                                                   '../data/raw/conll2003/en/train.tsv')
    token_counter_valid = conll2003.convert_to_tsv('../data/raw/conll2003/en/valid_bio.txt',
                                                   '../data/raw/conll2003/en/valid.tsv')
    conll2003.convert_to_tfrecord('../data/raw/conll2003/en/train.tsv', '../dataset/train/conll2003.tfrecords')
    conll2003.convert_to_tfrecord('../data/raw/conll2003/en/valid.tsv', '../dataset/valid/conll2003.tfrecords')

    xlsx = XlsxDataProcessor(Config())

    token_c1, token_c2 = xlsx.convert_to_tsv('../data/raw/City_biz_incubator/BusinessEcosystem.xlsx',
                                             '../data/raw/City_biz_incubator/train.tsv',
                                             '../data/raw/City_biz_incubator/valid.tsv')
    token_counter_train = Counter(token_counter_train)
    token_c1 = Counter(token_c1)
    token_counter_train = dict(token_counter_train + token_c1)
    token_counter_valid = Counter(token_counter_valid)
    token_c2 = Counter(token_c2)
    token_counter_valid = dict(token_counter_valid + token_c2)
    xlsx.convert_to_tfrecord('../data/raw/City_biz_incubator/train.tsv', '../dataset/train/City_biz_incubator_BusinessEcosystem.tfrecords')
    xlsx.convert_to_tfrecord('../data/raw/City_biz_incubator/valid.tsv',
                             '../dataset/valid/City_biz_incubator_BusinessEcosystem.tfrecords')

    # token_c1, token_c2 = xlsx.convert_to_tsv('../data/raw/City_door_open/Doors_Open_2018.xlsx',
    #                                          '../data/raw/City_door_open/train.tsv',
    #                                          '../data/raw/City_door_open/valid.tsv')
    # token_counter_train = Counter(token_counter_train)
    # token_c1 = Counter(token_c1)
    # token_counter_train = dict(token_counter_train + token_c1)
    # token_counter_valid = Counter(token_counter_valid)
    # token_c2 = Counter(token_c2)
    # token_counter_valid = dict(token_counter_valid + token_c2)
    # xlsx.convert_to_tfrecord('../data/raw/City_door_open/train.tsv',
    #                          '../dataset/train/City_door_open_2018.tfrecords')
    # xlsx.convert_to_tfrecord('../data/raw/City_door_open/valid.tsv',
    #                          '../dataset/valid/City_door_open_2018.tfrecords')

    print(">>>>> Training Set")
    for k in token_counter_train:
        print(k, ':\t\t', token_counter_train[k])

    print(">>>>> Evaluation Set")
    for k in token_counter_valid:
        print(k, ':\t\t', token_counter_valid[k])
